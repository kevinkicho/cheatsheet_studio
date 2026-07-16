#!/usr/bin/env python3
"""
Build Content Mastersheet from catalog + everything sheet + topic packs.

Outputs (regenerate anytime):
  packages/cheatsheet-sdk/data/content-mastersheet.xlsx
  packages/cheatsheet-sdk/data/content-inventory.json

Usage:
  python scripts/build-content-mastersheet.py
"""
from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
CATALOG = ROOT / "packages/cheatsheet-sdk/data/seed-catalog.json"
SHEET = ROOT / "examples/agent-out/everything.sheet.json"
PACKS_DIR = ROOT / "packages/cheatsheet-sdk/topic-packs"
PACKS_INDEX = PACKS_DIR / "index.json"
OUT_XLSX = ROOT / "packages/cheatsheet-sdk/data/content-mastersheet.xlsx"
OUT_JSON = ROOT / "packages/cheatsheet-sdk/data/content-inventory.json"

# Canonical outline for generation tracking (subject → topics → target card count)
# Aligns with kitchen-sink L1/L2 + expansion plan. Adjust targets over time.
OUTLINE_TARGETS: dict[str, dict[str, int]] = {
    "mathematics": {
        "Algebra": 12,
        "Trigonometry": 12,
        "Geometry": 10,
        "Precalculus": 8,
        "Calculus": 30,
        "Differential Equations": 12,
        "Linear Algebra": 16,
        "Complex Numbers": 8,
        "Sequences": 10,
        "Probability": 14,
        "Statistics": 14,
        "Combinatorics": 8,
        "Discrete Math": 10,
        "Number Theory": 6,
        "Notation": 6,
        "General": 4,
    },
    "physics": {
        "Constants": 8,
        "Mechanics": 24,
        "Waves": 8,
        "Optics": 8,
        "Electricity & Magnetism": 16,
        "Electromagnetism": 0,  # merge into E&M
        "Thermodynamics": 10,
        "Modern Physics": 10,
        "Fluids": 8,
        "Quantum": 6,
    },
    "chemistry": {
        "Stoichiometry": 10,
        "Gases": 8,
        "Acids & Bases": 12,
        "Aqueous Chemistry": 8,
        "Thermodynamics": 10,
        "Kinetics": 10,
        "Electrochemistry": 8,
        "Spectroscopy": 6,
        "Atomic Structure": 8,
        "Bonding": 8,
        "Organic": 12,
        "General": 4,
    },
    "biology": {
        "Biochemistry": 8,
        "Cell Biology": 8,
        "Molecular Biology": 10,
        "Genetics": 10,
        "Population Genetics": 6,
        "Metabolism": 8,
        "Enzymes": 8,
        "Physiology": 8,
        "Neurophysiology": 6,
        "Ecology": 8,
        "Plant Biology": 6,
        "Population": 4,
        "Immunology": 4,
    },
    "economics": {
        "Microeconomics": 20,
        "Macroeconomics": 16,
        "International": 6,
        "Public Economics": 4,
    },
    "finance": {
        "Time Value of Money": 12,
        "Capital Budgeting": 8,
        "Portfolio Theory": 10,
        "Asset Pricing": 8,
        "Fixed Income": 10,
        "Derivatives": 10,
        "Corporate Finance": 8,
        "Equity Valuation": 6,
        "Risk & Returns": 6,
        "Accounting": 6,
        "Analysis": 4,
    },
    "general": {
        "Templates": 4,
        "Study skills": 4,
    },
}

# Topic name aliases → canonical (case-insensitive match on topic)
TOPIC_ALIASES = {
    "asset pricing": "Asset Pricing",
    "capital budgeting": "Capital Budgeting",
    "corporate finance": "Corporate Finance",
    "electricity & magnetism": "Electricity & Magnetism",
    "electromagnetism": "Electricity & Magnetism",
    "sequences": "Sequences",
    "series": "Sequences",
}


def norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


def norm_subject(s: str) -> str:
    return norm_ws(s).lower() or "unknown"


def norm_topic(s: str) -> str:
    t = norm_ws(s)
    if not t:
        return "(unassigned)"
    key = t.lower()
    return TOPIC_ALIASES.get(key, t)


def latex_fingerprint(latex: str) -> str:
    if not latex:
        return ""
    s = latex.lower()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"\\left|\\right", "", s)
    s = re.sub(r"[{}]", "", s)
    return s[:200]


def content_key(title: str, latex: str, typ: str) -> str:
    fp = latex_fingerprint(latex)
    base = f"{typ}|{norm_ws(title).lower()}|{fp}"
    return hashlib.sha1(base.encode("utf-8")).hexdigest()[:12]


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def folder_maps(folders: list[dict]):
    by_id = {f["id"]: f for f in folders}
    def chain(fid: str | None) -> list[str]:
        names = []
        seen = set()
        while fid and fid in by_id and fid not in seen:
            seen.add(fid)
            f = by_id[fid]
            names.append(f.get("name") or fid)
            fid = f.get("parentId")
        names.reverse()
        return names
    return by_id, chain


def subject_from_folder_path(path: list[str]) -> str:
    if not path:
        return "unknown"
    top = path[0]
    # "1. Biology" → biology
    m = re.match(r"^\d+\.\s*(.+)$", top)
    name = (m.group(1) if m else top).strip().lower()
    return name


def topic_from_folder_path(path: list[str]) -> str:
    if len(path) < 2:
        return "(root)"
    leaf = path[-1]
    m = re.match(r"^\d+(?:\.\d+)*\s+(.+)$", leaf)
    return norm_topic(m.group(1) if m else leaf)


def collect_inventory():
    cat = load_json(CATALOG)
    sheet = load_json(SHEET) if SHEET.exists() else {"items": [], "folders": []}
    packs_idx = load_json(PACKS_INDEX) if PACKS_INDEX.exists() else {"packs": []}

    folders = sheet.get("folders") or []
    by_id, chain = folder_maps(folders)

    # content_key → record
    inv: dict[str, dict] = {}

    def upsert(rec: dict, source: str):
        ck = rec["content_key"]
        if ck not in inv:
            inv[ck] = {
                **rec,
                "in_catalog": False,
                "in_everything_sheet": False,
                "in_topic_packs": [],
                "catalog_id": "",
                "sheet_ids": [],
            }
        row = inv[ck]
        # Prefer non-empty fields
        for k in ("title", "type", "subject", "topic", "latex", "description", "tags"):
            if rec.get(k) and not row.get(k):
                row[k] = rec[k]
        if source == "catalog":
            row["in_catalog"] = True
            row["catalog_id"] = rec.get("catalog_id") or row["catalog_id"]
        elif source == "sheet":
            row["in_everything_sheet"] = True
            sid = rec.get("sheet_id")
            if sid and sid not in row["sheet_ids"]:
                row["sheet_ids"].append(sid)
        elif source.startswith("pack:"):
            pid = source.split(":", 1)[1]
            if pid not in row["in_topic_packs"]:
                row["in_topic_packs"].append(pid)

    # Catalog
    for it in cat.get("items") or []:
        title = it.get("title") or it.get("id") or ""
        latex = it.get("latex") or ""
        typ = it.get("type") or "equation"
        ck = content_key(title, latex, typ)
        upsert(
            {
                "content_key": ck,
                "catalog_id": it.get("id") or "",
                "type": typ,
                "title": title,
                "subject": norm_subject(it.get("subject") or ""),
                "topic": norm_topic(it.get("topic") or ""),
                "latex": latex,
                "description": it.get("description") or "",
                "tags": ", ".join(it.get("tags") or []),
            },
            "catalog",
        )

    # Everything sheet (skip pure heading banners if no useful latex title-only)
    for it in sheet.get("items") or []:
        if it.get("hidden"):
            continue
        typ = it.get("type") or "equation"
        title = it.get("title") or it.get("id") or ""
        latex = it.get("latex") or ""
        # Resolve subject/topic from folder path
        path = chain(it.get("folderId"))
        subject = subject_from_folder_path(path)
        topic = topic_from_folder_path(path)
        # Headings like "1. Biology" at root
        if re.match(r"^\d+\.\s+", title) and not latex:
            continue
        ck = content_key(title, latex, typ)
        upsert(
            {
                "content_key": ck,
                "sheet_id": it.get("id") or "",
                "type": typ,
                "title": title,
                "subject": subject,
                "topic": topic,
                "latex": latex,
                "description": "",
                "tags": "",
                "folder_path": " / ".join(path),
            },
            "sheet",
        )

    # Topic packs
    pack_rows = []
    for meta in packs_idx.get("packs") or []:
        pid = meta.get("id") or ""
        fpath = PACKS_DIR / (meta.get("file") or f"{pid}.json")
        if not fpath.exists():
            continue
        pack = load_json(fpath)
        outline = pack.get("outline") or {}
        blocks = outline.get("blocks") or []
        subjects = pack.get("subjects") or meta.get("subjects") or []
        block_count = 0
        catalog_refs = []
        for b in blocks:
            btype = b.get("type") or ""
            if btype == "heading":
                continue
            block_count += 1
            if btype == "catalog":
                catalog_refs.append(b.get("id") or "")
                # mark catalog item as in pack
                cid = b.get("id") or ""
                for row in inv.values():
                    if row.get("catalog_id") == cid:
                        if pid not in row["in_topic_packs"]:
                            row["in_topic_packs"].append(pid)
                continue
            title = b.get("title") or ""
            latex = b.get("latex") or ""
            typ = "equation" if btype == "equation" else btype
            ck = content_key(title, latex, typ)
            subj = norm_subject(subjects[0] if subjects else "unknown")
            upsert(
                {
                    "content_key": ck,
                    "type": typ,
                    "title": title,
                    "subject": subj,
                    "topic": "(pack-only)",
                    "latex": latex,
                    "description": pack.get("description") or "",
                    "tags": f"pack:{pid}",
                },
                f"pack:{pid}",
            )
        pack_rows.append(
            {
                "pack_id": pid,
                "title": pack.get("title") or meta.get("title") or pid,
                "file": meta.get("file") or "",
                "subjects": ", ".join(subjects),
                "description": pack.get("description") or meta.get("description") or "",
                "block_count": block_count,
                "catalog_refs": ", ".join(catalog_refs),
            }
        )

    # Folder hierarchy rows
    folder_rows = []
    for f in folders:
        path = chain(f.get("id"))
        folder_rows.append(
            {
                "folder_id": f.get("id"),
                "name": f.get("name"),
                "parent_id": f.get("parentId") or "",
                "order": f.get("order", 0),
                "path": " / ".join(path),
                "depth": len(path),
                "subject": subject_from_folder_path(path),
                "topic": topic_from_folder_path(path) if len(path) >= 2 else "(L1)",
            }
        )

    # Counts matrix
    counts: dict[tuple[str, str], dict] = defaultdict(
        lambda: {
            "catalog": 0,
            "sheet": 0,
            "either": 0,
            "pack_only": 0,
        }
    )
    for row in inv.values():
        key = (row["subject"], row["topic"])
        c = counts[key]
        if row["in_catalog"]:
            c["catalog"] += 1
        if row["in_everything_sheet"]:
            c["sheet"] += 1
        if row["in_catalog"] or row["in_everything_sheet"]:
            c["either"] += 1
        if row["in_topic_packs"] and not row["in_catalog"] and not row["in_everything_sheet"]:
            c["pack_only"] += 1

    # Generate plan vs targets
    plan_rows = []
    for subject, topics in OUTLINE_TARGETS.items():
        for topic, target in topics.items():
            c = counts.get((subject, topic), {"catalog": 0, "sheet": 0, "either": 0})
            have = c["either"]
            # also count catalog-only under alias
            gap = max(0, target - have) if target > 0 else 0
            status = (
                "merge/skip"
                if target == 0
                else "full"
                if gap == 0 and have > 0
                else "empty"
                if have == 0
                else "partial"
            )
            plan_rows.append(
                {
                    "subject": subject,
                    "topic": topic,
                    "have": have,
                    "catalog": c["catalog"],
                    "sheet": c["sheet"],
                    "target": target,
                    "gap": gap,
                    "status": status,
                    "priority": (
                        1
                        if status == "empty" and target >= 8
                        else 2
                        if status == "partial" and gap >= 4
                        else 3
                        if status in ("empty", "partial")
                        else 4
                    ),
                }
            )
    plan_rows.sort(key=lambda r: (r["priority"], -r["gap"], r["subject"], r["topic"]))

    # Duplicate candidates: same title (casefold) different content_key
    by_title: dict[str, list] = defaultdict(list)
    by_latex: dict[str, list] = defaultdict(list)
    for row in inv.values():
        t = norm_ws(row.get("title") or "").lower()
        if t:
            by_title[t].append(row)
        fp = latex_fingerprint(row.get("latex") or "")
        if fp and len(fp) > 8:
            by_latex[fp].append(row)

    dup_rows = []
    seen_pairs = set()
    for t, rows in by_title.items():
        if len(rows) < 2:
            continue
        ids = sorted({r.get("catalog_id") or r["content_key"] for r in rows})
        key = ("title", t)
        if key in seen_pairs:
            continue
        seen_pairs.add(key)
        dup_rows.append(
            {
                "kind": "same_title",
                "key": t,
                "count": len(rows),
                "ids": ", ".join(ids[:12]),
                "subjects": ", ".join(sorted({r["subject"] for r in rows})),
                "topics": ", ".join(sorted({r["topic"] for r in rows})),
            }
        )
    for fp, rows in by_latex.items():
        if len(rows) < 2:
            continue
        titles = {norm_ws(r.get("title") or "").lower() for r in rows}
        if len(titles) == 1:
            continue  # same title already listed
        ids = sorted({r.get("catalog_id") or r["content_key"] for r in rows})
        dup_rows.append(
            {
                "kind": "same_latex",
                "key": fp[:60],
                "count": len(rows),
                "ids": ", ".join(ids[:12]),
                "subjects": ", ".join(sorted({r["subject"] for r in rows})),
                "topics": ", ".join(sorted({r.get("title") or "" for r in rows})),
            }
        )
    dup_rows.sort(key=lambda r: (-r["count"], r["kind"], r["key"]))

    inventory = sorted(
        inv.values(),
        key=lambda r: (r["subject"], r["topic"], r.get("title") or ""),
    )

    meta = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "catalogCount": cat.get("count") or len(cat.get("items") or []),
        "catalogTypes": cat.get("types") or {},
        "sheetItems": len(sheet.get("items") or []),
        "sheetFolders": len(folders),
        "uniqueContentKeys": len(inventory),
        "topicPacks": len(pack_rows),
        "sources": {
            "catalog": str(CATALOG.relative_to(ROOT)),
            "everythingSheet": str(SHEET.relative_to(ROOT)) if SHEET.exists() else None,
            "topicPacks": str(PACKS_DIR.relative_to(ROOT)),
        },
    }

    return {
        "meta": meta,
        "inventory": inventory,
        "counts": counts,
        "folders": folder_rows,
        "packs": pack_rows,
        "plan": plan_rows,
        "duplicates": dup_rows,
    }


# ── Excel ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(name="Arial", bold=True, color="F8FAFC", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="0F172A")
SUB_FONT = Font(name="Arial", size=10, color="334155")
BODY_FONT = Font(name="Arial", size=10)
BLUE_INPUT = Font(name="Arial", size=10, color="0000FF")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
FILL_FULL = PatternFill("solid", fgColor="BBF7D0")
FILL_PARTIAL = PatternFill("solid", fgColor="FEF08A")
FILL_EMPTY = PatternFill("solid", fgColor="FECACA")
FILL_SKIP = PatternFill("solid", fgColor="E2E8F0")
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")
FILL_YELLOW = PatternFill("solid", fgColor="FEF9C3")


def style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[row].height = 28


def autosize(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 8
        for cell in col[:80]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def add_table(ws, name: str, ref: str):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(tab)


def write_overview(wb: Workbook, data: dict):
    ws = wb.active
    ws.title = "Overview"
    m = data["meta"]
    ws["A1"] = "CheatSheet Studio — Content Mastersheet"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Inventory of catalog + everything.sheet + topic packs. "
        "Use Generate_Plan to avoid duplicate generation. Re-run: "
        "python scripts/build-content-mastersheet.py"
    )
    ws["A2"].font = SUB_FONT
    ws.merge_cells("A2:F2")

    rows = [
        ("Generated (UTC)", m["generatedAt"]),
        ("Unique content keys", m["uniqueContentKeys"]),
        ("Catalog items", m["catalogCount"]),
        ("Catalog equations", m["catalogTypes"].get("equation", "")),
        ("Catalog tables", m["catalogTypes"].get("table", "")),
        ("Catalog figures", m["catalogTypes"].get("figure", "")),
        ("Catalog processes", m["catalogTypes"].get("process", "")),
        ("Everything sheet items", m["sheetItems"]),
        ("Everything sheet folders", m["sheetFolders"]),
        ("Topic packs", m["topicPacks"]),
        ("Catalog path", m["sources"]["catalog"]),
        ("Sheet path", m["sources"]["everythingSheet"]),
        ("Packs path", m["sources"]["topicPacks"]),
    ]
    ws["A4"] = "Snapshot"
    ws["A4"].font = Font(name="Arial", bold=True, size=12)
    for i, (k, v) in enumerate(rows, start=5):
        ws.cell(i, 1, k).font = Font(name="Arial", bold=True, size=10)
        ws.cell(i, 2, v).font = BODY_FONT

    ws["A20"] = "How to use"
    ws["A20"].font = Font(name="Arial", bold=True, size=12)
    tips = [
        "1. Inventory — every unique formula/table/process (deduped by title+latex fingerprint).",
        "2. Matrix — counts by subject × topic (catalog vs everything sheet).",
        "3. Generate_Plan — target vs have; gap>0 = safe to generate (status empty/partial).",
        "4. Before generating: filter Inventory by subject/topic; check Duplicates for same title/latex.",
        "5. After adding content: re-run this script so have/gap stay accurate.",
        "6. Blue cells on Generate_Plan (target) are inputs you may edit; then recompute gap = max(0, target-have).",
        "7. content_key is stable-ish dedupe id; catalog_id is seed-catalog id when present.",
    ]
    for i, t in enumerate(tips, start=21):
        ws.cell(i, 1, t).font = SUB_FONT
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

    ws["A30"] = "Subjects in inventory"
    ws["A30"].font = Font(name="Arial", bold=True, size=12)
    subj_counts: dict[str, int] = defaultdict(int)
    for r in data["inventory"]:
        subj_counts[r["subject"]] += 1
    ws["A31"] = "Subject"
    ws["B31"] = "Unique items"
    style_header(ws, 31, 2)
    for i, (s, n) in enumerate(sorted(subj_counts.items()), start=32):
        ws.cell(i, 1, s).font = BODY_FONT
        ws.cell(i, 2, n).font = BODY_FONT

    autosize(ws)


def write_inventory(wb: Workbook, data: dict):
    ws = wb.create_sheet("Inventory")
    headers = [
        "content_key",
        "catalog_id",
        "type",
        "title",
        "subject",
        "topic",
        "in_catalog",
        "in_everything_sheet",
        "in_topic_packs",
        "sheet_ids",
        "tags",
        "description",
        "latex",
        "latex_fingerprint",
        "folder_path",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))

    for r_i, row in enumerate(data["inventory"], start=2):
        packs = row.get("in_topic_packs") or []
        vals = [
            row.get("content_key"),
            row.get("catalog_id"),
            row.get("type"),
            row.get("title"),
            row.get("subject"),
            row.get("topic"),
            "Y" if row.get("in_catalog") else "N",
            "Y" if row.get("in_everything_sheet") else "N",
            ", ".join(packs),
            ", ".join(row.get("sheet_ids") or []),
            row.get("tags"),
            row.get("description"),
            row.get("latex"),
            latex_fingerprint(row.get("latex") or ""),
            row.get("folder_path") or "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_i, c, v)
            cell.font = BODY_FONT
            cell.border = THIN
            if r_i % 2 == 0:
                cell.fill = FILL_ALT
            if c in (7, 8) and v == "N":
                cell.fill = FILL_YELLOW

    last = max(2, len(data["inventory"]) + 1)
    add_table(ws, "InventoryTable", f"A1:{get_column_letter(len(headers))}{last}")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = "A2"
    autosize(ws)
    ws.column_dimensions["M"].width = 40
    ws.column_dimensions["N"].width = 24


def write_matrix(wb: Workbook, data: dict):
    ws = wb.create_sheet("Matrix")
    headers = [
        "subject",
        "topic",
        "have_either",
        "catalog",
        "everything_sheet",
        "pack_only_extra",
        "target",
        "gap",
        "status",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))

    # Build target lookup
    targets = {
        (s, t): tgt
        for s, topics in OUTLINE_TARGETS.items()
        for t, tgt in topics.items()
    }

    rows = []
    for (subject, topic), c in sorted(data["counts"].items()):
        target = targets.get((subject, topic), "")
        have = c["either"]
        gap = max(0, int(target) - have) if target != "" else ""
        if target == 0:
            status = "merge/skip"
        elif target == "":
            status = "unplanned"
        elif have == 0:
            status = "empty"
        elif gap == 0:
            status = "full"
        else:
            status = "partial"
        rows.append(
            (
                subject,
                topic,
                have,
                c["catalog"],
                c["sheet"],
                c["pack_only"],
                target if target != "" else None,
                gap if gap != "" else None,
                status,
            )
        )
    # Add planned topics with zero have
    seen = {(r[0], r[1]) for r in rows}
    for s, topics in OUTLINE_TARGETS.items():
        for t, target in topics.items():
            if (s, t) in seen:
                continue
            status = "merge/skip" if target == 0 else "empty"
            rows.append((s, t, 0, 0, 0, 0, target, max(0, target), status))
    rows.sort(key=lambda r: (r[0], r[1]))

    for r_i, row in enumerate(rows, start=2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r_i, c, v)
            cell.font = BODY_FONT
            cell.border = THIN
            if c == 7 and v is not None:
                cell.font = BLUE_INPUT
                cell.fill = FILL_YELLOW
        st = row[8]
        fill = {
            "full": FILL_FULL,
            "partial": FILL_PARTIAL,
            "empty": FILL_EMPTY,
            "merge/skip": FILL_SKIP,
            "unplanned": FILL_ALT,
        }.get(st)
        if fill:
            ws.cell(r_i, 9).fill = fill

    last = max(2, len(rows) + 1)
    add_table(ws, "MatrixTable", f"A1:I{last}")
    ws.freeze_panes = "A2"
    autosize(ws)


def write_plan(wb: Workbook, data: dict):
    ws = wb.create_sheet("Generate_Plan")
    ws["A1"] = "Generation queue — only create cards where gap > 0 and status ≠ full/merge"
    ws["A1"].font = SUB_FONT
    ws.merge_cells("A1:I1")

    headers = [
        "priority",
        "subject",
        "topic",
        "have",
        "catalog",
        "sheet",
        "target",
        "gap",
        "status",
        "suggested_action",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)
    style_header(ws, 2, len(headers))

    for r_i, row in enumerate(data["plan"], start=3):
        action = {
            "empty": "Generate full topic pack (core formulas first)",
            "partial": "Generate missing cards only; check Inventory for existing titles",
            "full": "Skip — at/above target",
            "merge/skip": "Do not generate; merge alias into canonical topic",
        }.get(row["status"], "")
        vals = [
            row["priority"],
            row["subject"],
            row["topic"],
            row["have"],
            row["catalog"],
            row["sheet"],
            row["target"],
            row["gap"],
            row["status"],
            action,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_i, c, v)
            cell.font = BODY_FONT
            cell.border = THIN
            if c == 7:
                cell.font = BLUE_INPUT
                cell.fill = FILL_YELLOW
        st = row["status"]
        fill = {
            "full": FILL_FULL,
            "partial": FILL_PARTIAL,
            "empty": FILL_EMPTY,
            "merge/skip": FILL_SKIP,
        }.get(st)
        if fill:
            ws.cell(r_i, 9).fill = fill

    last = max(3, len(data["plan"]) + 2)
    # gap formula optional for future edits: H = max(0, G-D)
    for r in range(3, last + 1):
        ws.cell(r, 8, f"=MAX(0,G{r}-D{r})")
        ws.cell(r, 8).font = BODY_FONT

    add_table(ws, "GeneratePlanTable", f"A2:J{last}")
    ws.freeze_panes = "A3"
    autosize(ws)
    ws.column_dimensions["J"].width = 55


def write_folders(wb: Workbook, data: dict):
    ws = wb.create_sheet("Folders_Everything")
    headers = [
        "folder_id",
        "name",
        "parent_id",
        "order",
        "depth",
        "subject",
        "topic",
        "path",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))
    rows = sorted(data["folders"], key=lambda r: (r["path"], r["order"]))
    for r_i, row in enumerate(rows, start=2):
        for c, k in enumerate(headers, 1):
            cell = ws.cell(r_i, c, row.get(k))
            cell.font = BODY_FONT
            cell.border = THIN
    last = max(2, len(rows) + 1)
    add_table(ws, "FoldersTable", f"A1:H{last}")
    ws.freeze_panes = "A2"
    autosize(ws)


def write_packs(wb: Workbook, data: dict):
    ws = wb.create_sheet("Topic_Packs")
    headers = [
        "pack_id",
        "title",
        "file",
        "subjects",
        "block_count",
        "catalog_refs",
        "description",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header(ws, 1, len(headers))
    for r_i, row in enumerate(data["packs"], start=2):
        for c, k in enumerate(headers, 1):
            cell = ws.cell(r_i, c, row.get(k))
            cell.font = BODY_FONT
            cell.border = THIN
    last = max(2, len(data["packs"]) + 1)
    add_table(ws, "PacksTable", f"A1:G{last}")
    ws.freeze_panes = "A2"
    autosize(ws)


def write_duplicates(wb: Workbook, data: dict):
    ws = wb.create_sheet("Duplicates")
    ws["A1"] = (
        "Possible duplicates — review before generating. "
        "same_title: identical titles; same_latex: similar latex different titles."
    )
    ws["A1"].font = SUB_FONT
    ws.merge_cells("A1:F1")
    headers = ["kind", "key", "count", "ids", "subjects", "topics_or_titles"]
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)
    style_header(ws, 2, len(headers))
    for r_i, row in enumerate(data["duplicates"], start=3):
        vals = [
            row["kind"],
            row["key"],
            row["count"],
            row["ids"],
            row["subjects"],
            row["topics"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_i, c, v)
            cell.font = BODY_FONT
            cell.border = THIN
            if row["count"] >= 3:
                cell.fill = FILL_PARTIAL
    last = max(3, len(data["duplicates"]) + 2)
    if data["duplicates"]:
        add_table(ws, "DuplicatesTable", f"A2:F{last}")
    ws.freeze_panes = "A3"
    autosize(ws)


def write_xlsx(data: dict):
    wb = Workbook()
    write_overview(wb, data)
    write_inventory(wb, data)
    write_matrix(wb, data)
    write_plan(wb, data)
    write_folders(wb, data)
    write_packs(wb, data)
    write_duplicates(wb, data)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def write_json(data: dict):
    # JSON-serializable counts keys
    out = {
        "meta": data["meta"],
        "inventory": data["inventory"],
        "matrix": [
            {
                "subject": s,
                "topic": t,
                **c,
            }
            for (s, t), c in sorted(data["counts"].items())
        ],
        "outlineTargets": OUTLINE_TARGETS,
        "generatePlan": data["plan"],
        "folders": data["folders"],
        "packs": data["packs"],
        "duplicates": data["duplicates"],
    }
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(out, indent=2), encoding="utf-8")


def main():
    data = collect_inventory()
    write_xlsx(data)
    write_json(data)
    m = data["meta"]
    print(f"Wrote {OUT_XLSX.relative_to(ROOT)}")
    print(f"Wrote {OUT_JSON.relative_to(ROOT)}")
    print(
        f"Unique keys={m['uniqueContentKeys']} catalog={m['catalogCount']} "
        f"sheet_items={m['sheetItems']} packs={m['topicPacks']}"
    )
    empty = sum(1 for r in data["plan"] if r["status"] == "empty")
    partial = sum(1 for r in data["plan"] if r["status"] == "partial")
    print(f"Generate plan: empty_topics={empty} partial_topics={partial} dups={len(data['duplicates'])}")


if __name__ == "__main__":
    main()
